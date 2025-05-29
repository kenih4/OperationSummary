#!/usr/bin/env python
# coding: utf-8

# In[1]:


#!/usr/bin/python
# -*- coding: utf-8 -*-

from mdaq import pymdaq_web
import datetime
import time
import pandas as pd
import sys
import codecs
import os
import binascii
import requests
from requests.exceptions import Timeout
import pandas as pd
import openpyxl
import subprocess
import copy
import libCom
import schedule
from operator import itemgetter


# In[2]:


db = pymdaq_web.db("srweb-dmz-03", port=8888)


# In[3]:


################################GUN HV OFF 時間の取得################################


# In[4]:


#期間内のGUN ON/OFF時間を取得する。最大31日未満。
def get_gun_status_long_time(time_beg, time_end):
    limit_day = 31
    time_list = []
    hv_list = []
    
    if (time_end-time_beg).days >= limit_day or (time_end-time_beg).days < 0:
        ret = -1
        print("DB GET Time ERROR")
    else:
        time_list = schedule.get_days_period_list(time_beg, time_end, 2)
        for i in range(len(time_list)):
            hv_list.extend(get_gun_status_db(time_list[i]["start"], time_list[i]["end"]))
            time.sleep(1)       
    return hv_list

#範囲時間内のGUNのステータスを取得
def get_gun_status_db(dt_beg, dt_end):
    hv_list =[]
    signal_name = 'xfel_rf_gun_ctrl/status'
    
    ret = check_db_get_time(dt_beg, dt_end)
    
    if ret == 0:   
        res = db.get_data(signal_name,dt_beg.strftime("%Y/%m/%d %H:%M:%S"),dt_end.strftime("%Y/%m/%d %H:%M:%S"))
        if db.status()==pymdaq_web.DB_OK:
            res.sort()
            hv_list = get_gun_off_time(res, 2)
            hv_list = merge_periods_list(hv_list)
            return hv_list
        else:
            print("ERR: %d  Msg: %s"%(db.status(),db.err_msg()))

            
#時間が3日未満か確認する
def check_db_get_time(dt_beg, dt_end):

    limit_day =  3
    
    if (dt_end-dt_beg).days >= limit_day or (dt_end-dt_beg).days < 0:
        ret = -1
        print("DB GET Time ERROR")
    else:
        ret = 0    
    return ret

#DBから取得したGUNのStatusのbitが切り替わった時のリストを生成する(0:off, 1:on)。時間は分単位で丸める。
def get_gun_off_time(sig_res, bit1):
    sig_res.sort()
    list = []
    dict = {}
    status = "ON"

    for i in range(len(sig_res)):
        if (sig_res[i][1] & (1 << bit1) == 0):
            if status == "ON":
                off_time = date_floor_minute(datetime.datetime.strptime(sig_res[i][0], "%Y/%m/%d %H:%M:%S.%f").replace(microsecond = 0),0)
                status = "OFF"
        else:
            if status == "OFF":
                on_time = date_floor_minute(datetime.datetime.strptime(sig_res[i][0], "%Y/%m/%d %H:%M:%S.%f").replace(microsecond = 0),1)
                status = "ON"
                
                td = on_time - off_time
                if td.total_seconds() >2:#2秒以下を排除
                    dict = {"start":off_time, "end":on_time}
                    list.append(dict)
            
        if (status == "OFF") & (i == (len(sig_res)-1)):
            on_time = date_floor_minute(datetime.datetime.strptime(sig_res[i][0], "%Y/%m/%d %H:%M:%S.%f").replace(microsecond = 0),1)
            dict = {"start":off_time, "end":on_time}
            print("dict: ", dict)
            list.append(dict)
    return list


# In[5]:


#on/off list重複する時間をまとめる
def merge_periods_list(list):
    list_bk =list.copy()
    list_tmp = []
    dict = {}
    
    for i in range(len(list_bk)):
        dict = {"flg":1,"time":list_bk[i]["start"]} #start time
        list_tmp.append(dict)
        dict = {"flg":-1,"time":list_bk[i]["end"]} #end time
        list_tmp.append(dict)
        
    sorted_list = sorted(list_tmp, key=itemgetter('time', 'flg'))
    total = 0
    for i in range(len(sorted_list)):
        total += sorted_list[i]["flg"]
        sorted_list[i]["Sum"] = total 

    Sum_bk = 0
    offOnList = []
    
    
    for i in range(len(sorted_list)):
        if(sorted_list[i]["Sum"] == 1 and Sum_bk == 0):
            off_time = sorted_list[i]["time"]
        if(sorted_list[i]["Sum"] == 0 and Sum_bk == 1):
            on_time = sorted_list[i]["time"]
            offOnList.append({"start":off_time, "end":on_time})
        Sum_bk = sorted_list[i]["Sum"]

    return  offOnList


# In[ ]:





# In[6]:


#日付のstrデータを+offset日した値にする。
def time_offset_days(time, offset_days):
    tmp = datetime.datetime.strptime(time, "%Y/%m/%d %H:%M:%S")
    tmp = tmp + datetime.timedelta(days=offset_days)
    str = tmp.strftime("%Y/%m/%d %H:%M:%S")
    return str


#時間の分を+offset分した値で分で丸める。
def date_floor_minute(time, offset):
    time = time + datetime.timedelta(minutes=offset)
    time = time.replace(second = 0, microsecond = 0)    
    return time


# In[7]:


#EXCELにGUN HV OFF時間を書き込む
def write_excel_gun_hv_time(list):
    wb = openpyxl.load_workbook(schedule.計画時間ファイル)    
    ws = wb.create_sheet()
    ws.title = "test"
    sheet_name = "GUN HV OFF"
    offset = 12
    
    for ws2 in wb.worksheets:
        if ws2.title.endswith(sheet_name):
            wb.remove(ws2)
            
    write_excel_gun_hv_time_bl(ws, list, 2)
    write_excel_gun_hv_time_bl(ws, list, 3)
    libCom.auto_sheet_width(ws)
    ws.title = sheet_name    
    wb.save(schedule.計画時間ファイル)
    
    
#EXCELにBLごとのGUN HV OFF時間を書き込む
def write_excel_gun_hv_time_bl(ws, list, bl_num):

    if bl_num == 2:
        key = "bl2_time"
        column_offset = 0
    elif bl_num == 3:
        key = "bl3_time"
        column_offset = 6
    else:
        return -1
              
    ws.cell(1,1+column_offset ,value = key)
    ws.cell(2,1+column_offset,value = 'GUN HV OFF')
    ws.cell(2,2+column_offset,value = 'GUN HV ON')
    ws.cell(2,3+column_offset,value = '理由') 
    ws.cell(2,4+column_offset,value = '時間')
    ws.cell(2,5+column_offset,value = '運転種別')
    
    bl_list = schedule.extract_list_specified_key(list, key, "ユーザー", -1)

    #配列ループ
    for i in range(0,len(bl_list)):
        #列にリストを書き込み
        # ws.cell(i+3,1+column_offset,value = bl_list[i]["start"]).number_format = 'yyyy/mm/dd hh:mm'
        # ws.cell(i+3,2+column_offset,value = bl_list[i]["end"]).number_format = 'yyyy/mm/dd hh:mm'
        ws.cell(i+3,1+column_offset,value = bl_list[i]["start"])
        ws.cell(i+3,2+column_offset,value = bl_list[i]["end"])
        dtime = bl_list[i]["end"] - bl_list[i]["start"]
        if dtime.total_seconds() > 1800:
           ws.cell(i+3,3+column_offset,value = "一時立ち入り") 
        ws.cell(i+3,4+column_offset,value = format_timedelta(dtime)).number_format = '[hh]:mm:ss'
        ws.cell(i+3,5+column_offset,value = bl_list[i][key])


# In[8]:


#deltatime形式を[hh:mm:ss]の文字列形式に変換する
def format_timedelta(timedelta):
  total_sec = timedelta.total_seconds()
  # hours
  hours = total_sec // 3600 
  # remaining seconds
  remain = total_sec - (hours * 3600)
  # minutes
  minutes = remain // 60
  # remaining seconds
  seconds = remain - (minutes * 60)

  # total time
  return '{:02}:{:02}:{:02}'.format(int(hours), int(minutes), int(seconds))


# In[9]:


#GUNのOFF/ON時間のリストにBL毎の運転種別を追加する
def check_operation_mode(gunlist,flg):
    
    list_tmp = []
    if flg == 1:#計画時間のEXCEL読み込み
        bl2_list = schedule.read_xcel_bl_operation_time("bl2")
        bl3_list = schedule.read_xcel_bl_operation_time("bl3")
    else:#BL集計ファイルのEXCEL読み込み
        bl2_list = schedule.read_xcel_bl_operation_time_2(2)
        bl3_list = schedule.read_xcel_bl_operation_time_2(3)

    for i in range(len(gunlist)):
        off_time = gunlist[i]["start"]
        on_time = gunlist[i]["end"]
        bl2_time = ""
        bl3_time = ""
        
        for j in range(len(bl2_list)):
            if off_time >= bl2_list[j]['start'] and off_time <= bl2_list[j]['end'] and bl2_list[j]['運転種別'] != "ユニット合計":
                if on_time > bl2_list[j]['end']:
                    bl2_time = "期間要分割"
                else:
                    bl2_time = bl2_list[j]['運転種別']
                    
        for k in range(len(bl3_list)):
            if off_time >= bl3_list[k]['start'] and off_time <= bl3_list[k]['end'] and bl3_list[k]['運転種別'] != "ユニット合計":
                if on_time > bl3_list[k]['end']:
                    bl3_time = "期間要分割"
                else:
                    bl3_time = bl3_list[k]['運転種別']            
        
        
        dict = {"start":gunlist[i]["start"],"end":gunlist[i]["end"],"bl2_time":bl2_time, "bl3_time":bl3_time}
        # print(dict)
        list_tmp.append(dict)
    return list_tmp
    


# In[10]:


#日時を指定してGUNのOFF/ON時間の出力する
def output_excel_gun_hvoff_time():
    bl_num = input("加速器を選択してください。  1:SCSS  デフォルトはSACLA >>>")
    if bl_num =="": 
        bl_num = 2 #　1以外ならなんでもSACLA
#    print("bl_num:",bl_num)
    

#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    with open("dt_beg.txt", mode='r', encoding="UTF-8") as f:
        buff_dt_beg = f.read()
    val = input("開始日時を入力してください。　(例)2021/11/1 10:00  デフォルトは「" + str(buff_dt_beg) + "」    >>>")
    if not val:
        dt_beg = datetime.datetime.strptime(buff_dt_beg, "%Y/%m/%d %H:%M")
    else:
        try:
            dt_beg = datetime.datetime.strptime(val, "%Y/%m/%d %H:%M")
            with open("dt_beg.txt","w") as o:
                o.write(val)
        except ValueError:
            print ("エラー：日時のフォーマットが正しくありません。")
            sys.exit()

    dt_end = dt_beg +  datetime.timedelta(days=14)
    val = input("終了日時を入力してください。　(例)2021/11/15 10:00   デフォルトは2週間後「" + str(dt_end) + "」です。    >>>")
    if val:
        try:
            dt_end = datetime.datetime.strptime(val, "%Y/%m/%d %H:%M")
        except ValueError:
            print ("エラー：日時のフォーマットが正しくありません。")
            sys.exit()
#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

        
        

    
    if int(bl_num) == 1:
        write_excel_planned_time_bl(1, dt_beg, dt_end)
    else: 
        flg = 1
        """
        val = input("DEBUG: 計画時間をDBから読み込みますか　yes(y) or no(n)　>>>")
        if val == "yes" or val == "y":
            print("計画時間をDBから読み込みます")
            flg = 1
        else:
            print("計画時間をBL集計ファイルから読み込みます") 
            flg = 0
        """
        schedule.output_excel_planned_time(dt_beg, dt_end)
        gun_list = get_gun_status_long_time(dt_beg, dt_end)
        gun_list = check_operation_mode(gun_list, flg)
        libCom.print_list(gun_list)

        write_excel_gun_hv_time(gun_list)
    return 0


# In[11]:


#  start_time = datetime.datetime(2022, 6, 13, 10, 0)
#  end_time = datetime.datetime(2022, 6, 27, 22, 0)

#  libCom.print_list(get_gun_status_long_time(start_time, end_time))


#  ret = output_excel_gun_hvoff_time()
#  if ret == 0:
#      print("正常終了しました")
#      EXCEL = r"C:\Users\hasegawa-t\Desktop\OperationSummary\計画時間.xlsx"
#      subprocess.Popen(['start', EXCEL], shell=True)
#  else:
#      print("異常終了しました")
#  time.sleep(60)


# In[ ]:




