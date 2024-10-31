#!/usr/bin/env python
# coding: utf-8

# In[1]:


#!/usr/bin/python
# -*- coding: utf-8 -*-

import datetime
import time
import pandas as pd
import sys
import codecs
import os
import sys
from icalendar import Calendar, Event
import icalendar
import binascii
import requests
from requests.exceptions import Timeout
import pandas as pd
import openpyxl
import subprocess
import copy
from operator import itemgetter


# In[2]:


#グローバル変数
#計画時間ファイル = r"C:\Users\kenichi\Documents\OperationSummary\Modified\計画時間_Modified.xlsm" #ファイルが破損する
#計画時間ファイル = r"\\saclaopr18.spring8.or.jp\common\運転状況集計\最新\計画時間.xlsm" #ファイルが破損する
計画時間ファイル = r"\\saclaopr18.spring8.or.jp\common\運転状況集計\最新\計画時間.xlsx" #MOTO
BL1集計ファイル = r"\\saclaopr18.spring8.or.jp\common\運転状況集計\最新\SCSS\SCSS運転状況集計BL1.xlsm"
SCSS集計ファイル = r"\\saclaopr18.spring8.or.jp\common\運転状況集計\最新\SCSS\SCSS運転集計記録.xlsm"
BL2集計ファイル = r"\\saclaopr18.spring8.or.jp\common\運転状況集計\最新\SACLA\SACLA運転状況集計BL2.xlsm"
BL3集計ファイル = r"\\saclaopr18.spring8.or.jp\common\運転状況集計\最新\SACLA\SACLA運転状況集計BL3.xlsm"
SACLA集計ファイル = r"\\saclaopr18.spring8.or.jp\common\運転状況集計\最新\SACLA\SACLA運転集計記録.xlsm"


# In[3]:


#*************************スケジュールのリスト取得、操作*************************


# In[4]:


# iCalファイルよりEventを読み取り、リストに格納する
def read_ical_data(icalData):
 
    # Eventリスト（戻り値）
    list_schedules = []
 
    # icalファイル「テキスト」を解析し cal に取り込む
    cal = icalendar.Calendar.from_ical(icalData)
 
    for e in cal.walk():
        if e.name == 'VEVENT' :
            # Eventを1つずつ、辞書形式dict_scheduleに一旦代入し、それをリストlist_schedulesに追加する
            dict_schedule = {"title":str(e.decoded("summary"),'utf8') if e.get("summary") else "",
                            "place":str(e.decoded("location"),'utf8') if e.get("location") else "",
                            "desc":str(e.decoded("description"),'utf8') if e.get("description") else "",
                            "start":e.decoded("dtstart"),
                            "end":e.decoded("dtend"),
                            "updated":e.decoded("dtstamp")
                            }
            list_schedules.append(dict_schedule)
            #print(dict_schedule)
 
    # 予定表Eventを格納したリストを返す
    list_schedules = time_tzinfo_non_(list_schedules)
    return list_schedules

#urlからicalファイルのtextを取得する
def get_ical_data(url):

    #print(url)
    try:
        res = requests.get(url, timeout=(30.0,30.0))   
    except Exception as e:
        #print('Exception!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!@get_acc_sync	' + url)
        print(e.args)
        return ''
    else:
        res.raise_for_status()
        return res.text

#リストのタイムゾーンを消す
def time_tzinfo_non_(list):
    list_tmp =[]
    for i in range(len(list)):
        dict = list[i]        
 
        dict["start"] = tzinfo_delete(list[i]["start"])
        dict["end"] = tzinfo_delete(list[i]["end"])
        dict["updated"] = tzinfo_delete(list[i]["updated"])
        list_tmp.append(dict)
    list_tmp = sorted(list_tmp, key=lambda s: s['start'])
    return list


#date型をdatetime型に変換してタイムゾーンを消す
def tzinfo_delete(d):
    struct_time = d.timetuple()  
    return datetime.datetime(*struct_time[:6])


# In[5]:


#スケジュールのユーザー時間を抽出したリストを返す
def get_user_list(bl_num):
    
    if bl_num == 2:
        url = "http://calsrv01.spring8.or.jp/davical/public.php/plan/BL2?ticket=W67M6dJF"
    elif bl_num == 3:
        url = "http://calsrv01.spring8.or.jp/davical/public.php/plan/BL3?ticket=oGtBR6Qy"
    elif bl_num == 1:
        url = "http://calsrv01.spring8.or.jp/davical/public.php/plan/BL1?ticket=4BY9wtIG"
    elif bl_num == 0:
        url = "http://calsrv01.spring8.or.jp/davical/public.php/plan/tuning?ticket=sLLBQudt"
    else:
        return -1
    
    cal = get_ical_data(url)
    list_schedules = read_ical_data(cal)

    user_list = []
    for i in range(len(list_schedules)):                  
        user = ""
        title = list_schedules[i]['title']
        if bl_num == 0:
            dict = {"運転種別":"施設調整", "start":list_schedules[i]['start'],"end":list_schedules[i]['end'],"備考":title}
            user_list.append(dict)
        elif ("BL-study" in title or "BL study" in title): #2024/10/18追加　　こんなのがあり、Gがあるせいでユーザー運転と認識されてしまっていたため。「BL-study 10keV/30Hz（富樫、FROG）」
            pass
        elif "FCBT" in title or "G" in title:
            if "FCBT" in title:
                user = "FCBT"
            elif "G" in title:
                if title.startswith("G"):
                    title_slist = list_schedules[i]['title'].split()
                    user = title_slist[0]
                else:
                    title_slist = list_schedules[i]['title'].split('G', 1)
                    user = title_slist[0] + "G"
            # list = ["ユーザー", list_schedules[i]['start'], list_schedules[i]['end'], user]
            dict = {"運転種別":"ユーザー", "start":list_schedules[i]['start'],"end":list_schedules[i]['end'],"備考":user}
            user_list.append(dict)
            
        # print(list_schedules[i]['start'].strftime('%Y/%m/%d %H:%M') + "-" + list_schedules[i]['end'].strftime('%Y/%m/%d %H:%M') + " " + user)
    # print(user_list)
    return user_list


# In[6]:


#ユーザー時間の間を利用調整で埋めたリストを返す
def compensate_tuning_time(list, dt_beg, dt_end):
    list_tmp = copy.copy(list)  
    for i in range(len(list)):
        if  i == 0:
            if list[i]['start'] != dt_beg:
                dict = {"運転種別":"利用調整", "start":dt_beg,"end":list[i]['start'],"備考":""}
                list_tmp.append(dict)
        elif list[i]['start'] != list[i-1]['end'] and i != 0:
            dict = {"運転種別":"利用調整", "start":list[i-1]['end'],"end":list[i]['start'],"備考":""}
            list_tmp.append(dict)
    if len(list) != 0:
        if list[len(list)-1]['end'] != dt_end:
            dict = {"運転種別":"利用調整", "start":list[len(list)-1]['end'],"end":dt_end,"備考":""}
            list_tmp.append(dict)
    else:
        dict = {"運転種別":"利用調整", "start":dt_beg,"end":dt_end,"備考":""}
        list_tmp.append(dict)
    list_tmp = sorted(list_tmp, key=lambda s: s['start'])
    dict = {"運転種別":"ユニット合計", "start":dt_beg,"end":dt_end,"備考":""}
    list_tmp.append(dict)
    return list_tmp


# In[7]:


#*************************計画時間のEXCELに書き読み込みする*************************


# In[8]:


#指定したBLのスケジュールを計画時間EXCELに書き込む
def write_excel_planned_time_bl(bl_num, dt_beg, dt_end):
    list = get_list_period_start_time(get_user_list(bl_num), dt_beg, dt_end)
    #list = list.append(get_facility_list())
    list = sorted(list, key=lambda s: s['start'])

    if bl_num != 1:
        list.extend(get_list_period_start_time(get_user_list(0), dt_beg, dt_end))
        list = sorted(list, key=lambda s: s['start'])
        list = compensate_tuning_time(list, dt_beg, dt_end)
    
    sheet_name = "bl" + str(bl_num)
    write_excel_planned_time_sheet(list, sheet_name)

#計画時間EXCELに計画時間を書き込む
def write_excel_planned_time_sheet(list, sheet_name):
    wb = openpyxl.load_workbook(計画時間ファイル)    
    ws = wb.create_sheet()
    ws.title = "test"
    
    for ws2 in wb.worksheets:
        if ws2.title.endswith(sheet_name):
            wb.remove(ws2)
    
    ws.cell(1,1,value = '運転種別')
    ws.cell(1,2,value = 'start')
    ws.cell(1,3,value = 'end')
    ws.cell(1,4,value = '備考')

    for i in range(0,len(list)):

        #列にリストを書き込み
        ws.cell(i+2,1,value = list[i]['運転種別'])
        # ws.cell(i+2,2,value = list[i]['start'].strftime("%Y/%m/%d %H:%M"))
        # ws.cell(i+2,3,value = list[i]['end'].strftime("%Y/%m/%d %H:%M"))
        ws.cell(i+2,2,value = list[i]['start'])
        ws.cell(i+2,3,value = list[i]['end'])
        ws.cell(i+2,4,value = list[i]['備考'])
    auto_sheet_width(ws)
    ws.title = sheet_name    
    wb.save(計画時間ファイル)


# In[9]:


def auto_sheet_width(ws):
    for col in ws.columns:
        max_length = 0
        column = col[0].column

        for cell in col:
            if get_east_asian_width_count(str(cell.value)) > max_length:
                max_length = get_east_asian_width_count(str(cell.value))

        adjusted_width = (max_length + 2) * 1.0
        ws.column_dimensions[col[0].column_letter].width = adjusted_width
        
        
#半角を1文字、全角2を文字としてカウントする
import unicodedata
def get_east_asian_width_count(text):
    count = 0
    for c in text:
        if unicodedata.east_asian_width(c) in 'FWA':
            count += 2
        else:
            count += 1
    return count


# In[10]:


#BL2とBL3を計画時間EXCELに計画時間を書き込む
def output_excel_planned_time(dt_beg, dt_end):
    write_excel_planned_time_bl(2, dt_beg, dt_end)
    write_excel_planned_time_bl(3, dt_beg, dt_end)


# In[11]:


#計画時間EXCELからBLのスケジュールを取得し、時間をdatetimeへ変換する
# def get_bl_operation_time(bl):
#     list = read_xcel_bl_operation_time(bl)
#     list_key_strptime(list, "start", "%Y/%m/%d %H:%M")
#     list_key_strptime(list, "end", "%Y/%m/%d %H:%M")
#     return list

#計画時間EXCELからBLのスケジュールを読み込む
def read_xcel_bl_operation_time(sheet_name):
    # エクセルファイルの取り込み
    wb = openpyxl.load_workbook(計画時間ファイル)
    ws = wb[sheet_name]
    # １行目（列名のセル）
    header_cells = ws[1]
    # ２行目以降（データ）
    bl_list = []
    for row in ws.iter_rows(min_row=2):
        row_dic = {}
        # セルの値を「key-value」で登録
        for k, v in zip(header_cells, row):
            row_dic[k.value] = v.value
        bl_list.append(row_dic)
        
    return bl_list



# In[12]:


#日時を指定して計画時間のEXCELにスケジュールを出力する
def output_excel_schedule_time():
    bl_num = input("加速器を選択してください。  1:SCSS, Other:SACLA >>>")
      
    val = input("開始日時を入力してください。　(例)2021/11/1 10:00 >>>")
    try:
        dt_beg = datetime.datetime.strptime(val, "%Y/%m/%d %H:%M")
    except ValueError:
        print ("エラー：日時のフォーマットが正しくありません。")
        return -1

    val = input("終了日時を入力してください　(例)2021/11/15 10:00 >>>")
    try:
        dt_end = datetime.datetime.strptime(val, "%Y/%m/%d %H:%M")
    except ValueError:
        print ("エラー：日時のフォーマットが正しくありません。")
        return -1
    
    
    if int(bl_num) == 1:
        write_excel_planned_time_bl(1, dt_beg, dt_end)
    else: 
        val = input("計画時間を更新しますか　yes(y) or no(n)　>>>")
        if val == "yes" or val == "y":
            print("計画時間を更新します")
            output_excel_planned_time(dt_beg, dt_end)
        else:
            print("計画時間は更新しません") 

    return 0


# In[13]:


#*************************BL集計ファイルのEXCELから読み込む*************************


# In[14]:


#BL集計記録のEXCELからBLのスケジュールを読み込む
def read_xcel_bl_operation_time_2(bl_num):
    # エクセルファイルの取り込み
    
    if bl_num == 1:
        path = BL1集計ファイル
    elif bl_num == 2:
        path = BL2集計ファイル
    elif bl_num == 3:
        path = BL3集計ファイル
    else:
        return -1
    
    wb = openpyxl.load_workbook(path)
    ws = wb["運転予定時間"]

    header_cells = ws[3]
    bl_list = []
    for row in ws.iter_rows(min_row=4,max_col=4):
        row_dic = {}
        # セルの値を「key-value」で登録
        for k, v in zip(header_cells, row):
            row_dic[k.value] = v.value
        if row_dic["運転種別"] != None:
            row_dic.pop('ユニット')
            bl_list.append(row_dic)

        
    return bl_list


# In[15]:


#EXCELからシートの集計記録を読み込む
def read_xcel_fault_time(acc, sheet):#acc 1:SCSS, Other:SACLA
    # エクセルファイルの取り込み
    if acc == 1:
        path = SCSS集計ファイル
    else:
        path = SACLA集計ファイル
    
    wb = openpyxl.load_workbook(path)
    ws = wb[sheet]
    # １行目（列名のセル）
    header_cells = ws[2]
    # ２行目以降（データ）
    list = []
    for row in ws.iter_rows(min_row=3):
        row_dic = {}
        row_dic["運転種別"] = sheet
        # セルの値を「key-value」で登録
        for k, v in zip(header_cells, row):
            row_dic[k.value] = v.value
        if row_dic["start"] != None:
            if row_dic["BL"] != "BL2" and row_dic["BL"] != "BL3":
                row_dic["BL"] = "ALL"
            list.append(row_dic)
        
    return list


# In[16]:


#*************************開始時間・終了時間のdictのリストの操作*************************


# In[17]:


#開始時間が時間範囲内を抽出したリストを返す
def get_list_period_start_time(list, dt_beg, dt_end):
    list_tmp = []
        
    for i in range(len(list)):       
        if list[i]['start'] >= dt_beg and list[i]['start'] < dt_end:
            list_tmp.append(list[i])
    return list_tmp


#時間範囲内を抽出したリストを返す
def get_list_period_time(list, dt_beg, dt_end):
    list_tmp = []
        
    for i in range(len(list)):
        if (dt_beg <= list[i]["end"]) and (dt_end >= list[i]['start']):#期間内判定
       # if list[i]['start'] >= dt_beg and list[i]['start'] < dt_end:
            list_tmp.append(list[i])
    return list_tmp


# In[18]:


#listの指定したkeyのformat形式の時間文字列をdatetimeに変換する
def list_key_strptime(list, key, format):
    for i in range(len(list)):
        list[i][key] = datetime.datetime.strptime(list[i][key], format)      
    return list


# In[19]:


#辞書型のリストから指定したkeyと一致（一致 flg = 0, 不一致　flg = 0以外）するリストを返す
def extract_list_specified_key(list, key, find_value, flg):
    list_tmp = []
        
    for i in range(len(list)):
        if flg == 0:
            if list[i][key] == find_value:
                list_tmp.append(list[i]) 
        else:
            if list[i][key] != find_value:
                list_tmp.append(list[i]) 

    return list_tmp


# In[20]:


#期間を指定した日数毎に区切ったList取得する。
def get_days_period_list(start_time, end_time, days_period):
    time_list = []
    period_list = []

    
    if (end_time-start_time).days < 0:
        ret = -1
        print("Time ERROR")
    else:
        time_bk = start_time
        while (time_bk - end_time).days < 0:
            time_list.append(time_bk)
            time_bk = time_bk + datetime.timedelta(days = days_period)

        time_list.append(end_time)
        
        for i in range(len(time_list)):
            if i < (len(time_list)-1):
                 period_list.append({"start":time_list[i], "end":time_list[i+1]})
    
    return period_list


# In[ ]:





# In[21]:


# ret = output_excel_schedule_time()
# if ret == 0:
#     print("正常終了しました")
#     EXCEL = 計画時間ファイル
#     subprocess.Popen(['start', EXCEL], shell=True)
# else:
#     print("異常終了しました")
# time.sleep(60)


          

